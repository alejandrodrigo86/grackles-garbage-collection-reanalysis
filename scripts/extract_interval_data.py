"""
Script: extract_interval_data.py
Pipeline stage: 1. Data reconstruction
Analytical purpose: Parse the BORIS project, observation-time workbook, and aggregated state
exports; pair state starts and stops; reconstruct complete 30-minute observation intervals; and
produce quality-control totals.
Inputs: Observations/Grackles_AROD.v7.0.boris; Observations/Obs_Time.xlsx;
Observations/xlsx_aggregated/*.xlsx
Outputs: .codex_work/issue4/interval_data.json
Run-order position: 01
Key scientific assumption: Every media file was watched completely. Camera time is the exposure
denominator; events within a day are not treated as independent days.
Provenance note: This annotated copy preserves the executed analytical statements. Only
explanatory comments were added; see MANIFEST.csv for original and packaged hashes.
"""

import hashlib
import json
import math
import os
import re
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


# --- Project locations and reproducible configuration ---
ROOT = Path.cwd()
OBS_ROOT = ROOT / "Observations"
BORIS_PATH = OBS_ROOT / "Grackles_AROD.v7.0.boris"
OBS_TIME_PATH = OBS_ROOT / "Obs_Time.xlsx"
AGG_ROOT = OBS_ROOT / "xlsx_aggregated"
OUTPUT_PATH = Path(
    os.environ.get(
        "GRACKLES_INTERVAL_OUTPUT",
        str(ROOT / ".codex_work" / "issue4" / "interval_data.json"),
    )
)

BIN_SECONDS = int(os.environ.get("GRACKLES_BIN_SECONDS", str(30 * 60)))
CONTAINMENT_TOLERANCE_SECONDS = 0.01
STATE_BEHAVIORS = {"At the observation site", "Searching for food"}
SUBJECT_UPDATE_BEHAVIORS = {
    "One subject",
    "Two subjects",
    "Three subjects",
    "Four subjects",
    "Five subjects",
    "Six subjects",
    "Seven subjects",
    "Eight subjects",
    "More subjects",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def numeric(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def observation_number(observation_id: str) -> int:
    return int(re.search(r"\d+", observation_id).group())


def parse_study_date(value: str):
    value = str(value)
    if value.endswith("2018"):
        return datetime.strptime(value, "%d/%m/%Y").date()
    if len(value.split("/")[-1]) == 4:
        return datetime.strptime(value, "%m/%d/%Y").date()
    return datetime.strptime(value, "%m/%d/%y").date()


def union_intervals(intervals):
    """Strict interval union: no unobserved gaps are filled."""
    ordered = sorted((float(start), float(stop)) for start, stop in intervals if stop >= start)
    merged = []
    for start, stop in ordered:
        if merged and start <= merged[-1][1]:
            merged[-1] = (merged[-1][0], max(merged[-1][1], stop))
        else:
            merged.append((start, stop))
    return merged


def interval_length(intervals):
    return sum(stop - start for start, stop in intervals)


def overlap_seconds(intervals, window_start, window_stop):
    return sum(
        max(0.0, min(stop, window_stop) - max(start, window_start))
        for start, stop in intervals
    )


def start_count(intervals, window_start, window_stop, include_stop=False):
    if include_stop:
        return sum(window_start <= start <= window_stop for start, _ in intervals)
    return sum(window_start <= start < window_stop for start, _ in intervals)


def iso_datetime(study_date, seconds_from_midnight):
    base = datetime.combine(study_date, datetime.min.time())
    return (base + timedelta(seconds=float(seconds_from_midnight))).isoformat(timespec="milliseconds")


def load_schedule():
    workbook = load_workbook(OBS_TIME_PATH, data_only=True, read_only=True)
    worksheet = workbook.active
    schedule = {}
    observer_prefix = {"Observer 1": "M", "Observer 2": "N", "Observer 3": "V"}
    for record_date, start_time, observer, start_seconds, classification in worksheet.iter_rows(
        min_row=2, values_only=True
    ):
        if not record_date or not observer or not classification:
            continue
        prefix = observer_prefix[observer]
        schedule[(prefix, int(classification))] = {
            "date": parse_study_date(record_date),
            "start_seconds": float(start_seconds),
            "observer": observer,
            "recorded_start_time": start_time.isoformat() if start_time else None,
        }
    return schedule


def load_state_intervals(observation_id, time_offset, media_duration):
    path = AGG_ROOT / f"{observation_id}.xlsx"
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    intervals = defaultdict(list)
    invalid = Counter()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[12] != "STATE" or row[9] not in STATE_BEHAVIORS:
            continue
        start = numeric(row[13])
        stop = numeric(row[14])
        if start is None or stop is None:
            invalid["missing_boundary"] += 1
            continue
        if time_offset:
            start -= time_offset
            stop -= time_offset
        if stop < start:
            invalid["negative_duration"] += 1
        if stop == start:
            invalid["zero_duration"] += 1
        if start < -CONTAINMENT_TOLERANCE_SECONDS or stop > media_duration + CONTAINMENT_TOLERANCE_SECONDS:
            invalid["outside_media"] += 1
        intervals[row[9]].append((start, stop))
    return intervals, invalid


def containment_audit(foraging_intervals, at_site_intervals):
    failures = []
    for start, stop in foraging_intervals:
        contained = any(
            start >= outer_start - CONTAINMENT_TOLERANCE_SECONDS
            and stop <= outer_stop + CONTAINMENT_TOLERANCE_SECONDS
            for outer_start, outer_stop in at_site_intervals
        )
        if not contained:
            failures.append((start, stop))
    return failures


def same_behavior_overlap_count(intervals):
    ordered = sorted(intervals)
    return sum(
        1
        for (_, previous_stop), (next_start, _) in zip(ordered, ordered[1:])
        if next_start < previous_stop - CONTAINMENT_TOLERANCE_SECONDS
    )


def subtract_covered_seconds(target_intervals, covering_intervals):
    target = union_intervals(target_intervals)
    covering = union_intervals(covering_intervals)
    return sum(
        (stop - start) - overlap_seconds(covering, start, stop)
        for start, stop in target
    )


def point_event_counts(events, window_start, window_stop, include_stop=False):
    counts = Counter()
    for event_time, behavior, modifier in events:
        in_window = window_start <= event_time <= window_stop if include_stop else window_start <= event_time < window_stop
        if not in_window:
            continue
        if behavior == "Fly":
            counts["fly_events"] += 1
        elif behavior == "Jump":
            counts["jump_events"] += 1
        elif behavior == "Out of the sight":
            counts["out_of_sight_events"] += 1
        elif behavior == "Garbage":
            counts["garbage_marker_events"] += 1
        elif behavior in SUBJECT_UPDATE_BEHAVIORS:
            counts["subject_count_update_events"] += 1
        elif behavior == "Displacement":
            counts["displacement_events"] += 1
            if modifier == "Displacement from focal to others":
                counts["displacement_focal_to_others"] += 1
            elif modifier == "Displacement from others to focal":
                counts["displacement_others_to_focal"] += 1
            else:
                counts["displacement_unspecified"] += 1
    return counts


# --- Execute the complete script workflow ---
def main():
    schedule = load_schedule()
    boris = json.loads(BORIS_PATH.read_text(encoding="utf-8"))

    interval_rows = []
    qc_rows = []
    day_rows = []

    for observation_id, observation in boris["observations"].items():
        prefix = observation_id[0]
        number = observation_number(observation_id)
        schedule_row = schedule[(prefix, number)]
        study_date = schedule_row["date"]
        camera_start_clock = schedule_row["start_seconds"]
        observer = schedule_row["observer"]

        media_duration = float(next(iter(observation["media_info"]["length"].values())))
        camera_end_clock = camera_start_clock + media_duration
        time_offset = float(observation.get("time offset", 0) or 0)

        state_intervals, invalid = load_state_intervals(observation_id, time_offset, media_duration)
        at_site_intervals = sorted(state_intervals["At the observation site"])
        foraging_intervals = sorted(state_intervals["Searching for food"])
        at_site_union = union_intervals(at_site_intervals)
        foraging_union = union_intervals(foraging_intervals)
        focal_visible_union = union_intervals(at_site_intervals + foraging_intervals)
        containment_failures = containment_audit(foraging_intervals, at_site_intervals)
        extra_visible_seconds = subtract_covered_seconds(foraging_intervals, at_site_intervals)

        normalized_events = []
        garbage_times = []
        for raw_event in observation["events"]:
            event_time = float(raw_event[0]) - time_offset if time_offset else float(raw_event[0])
            behavior = raw_event[2]
            modifier = raw_event[3]
            if behavior == "Garbage":
                garbage_times.append(event_time)
            if behavior not in STATE_BEHAVIORS:
                normalized_events.append((event_time, behavior, modifier))

        collection_day = bool(garbage_times)
        collection_end_relative = garbage_times[0] if collection_day else None
        collection_end_clock = camera_start_clock + collection_end_relative if collection_day else None

        first_clock_bin = math.floor(camera_start_clock / BIN_SECONDS) * BIN_SECONDS
        final_clock_bin = math.ceil(camera_end_clock / BIN_SECONDS) * BIN_SECONDS
        bin_number = 0
        current_clock_start = first_clock_bin
        while current_clock_start < final_clock_bin:
            current_clock_stop = current_clock_start + BIN_SECONDS
            relative_start = max(0.0, current_clock_start - camera_start_clock)
            relative_stop = min(media_duration, current_clock_stop - camera_start_clock)
            camera_seconds = max(0.0, relative_stop - relative_start)
            if camera_seconds <= 0:
                current_clock_start = current_clock_stop
                continue

            bin_number += 1
            is_final_bin = current_clock_stop >= final_clock_bin
            point_counts = point_event_counts(
                normalized_events, relative_start, relative_stop, include_stop=is_final_bin
            )

            raw_at_site_seconds = overlap_seconds(at_site_union, relative_start, relative_stop)
            foraging_seconds = overlap_seconds(foraging_union, relative_start, relative_stop)
            focal_visible_seconds = overlap_seconds(focal_visible_union, relative_start, relative_stop)
            added_visible_seconds = max(0.0, focal_visible_seconds - raw_at_site_seconds)

            if not collection_day:
                collection_phase = "No collection"
                post_collection_seconds = 0.0
                time_from_collection_midpoint_hours = None
            else:
                post_collection_seconds = max(
                    0.0, relative_stop - max(relative_start, collection_end_relative)
                )
                if relative_stop <= collection_end_relative:
                    collection_phase = "Pre-collection"
                elif relative_start >= collection_end_relative:
                    collection_phase = "Post-collection"
                else:
                    collection_phase = "Straddles collection end"
                observed_midpoint = (relative_start + relative_stop) / 2.0
                time_from_collection_midpoint_hours = (
                    observed_midpoint - collection_end_relative
                ) / 3600.0

            interval_rows.append(
                {
                    "observation_id": observation_id,
                    "observation_date": study_date.isoformat(),
                    "weekday": study_date.strftime("%A"),
                    "observer": observer,
                    "collection_day": int(collection_day),
                    "camera_start_datetime": iso_datetime(study_date, camera_start_clock),
                    "camera_end_datetime": iso_datetime(study_date, camera_end_clock),
                    "bin_number": bin_number,
                    "bin_start_datetime": iso_datetime(study_date, current_clock_start),
                    "bin_end_datetime": iso_datetime(study_date, current_clock_stop),
                    "camera_seconds": camera_seconds,
                    "collection_end_datetime": iso_datetime(study_date, collection_end_clock)
                    if collection_day
                    else None,
                    "collection_phase": collection_phase,
                    "post_collection_camera_seconds": post_collection_seconds,
                    "time_from_collection_midpoint_hours": time_from_collection_midpoint_hours,
                    "raw_at_site_seconds": raw_at_site_seconds,
                    "foraging_seconds": foraging_seconds,
                    "focal_visible_seconds": focal_visible_seconds,
                    "added_visible_seconds": added_visible_seconds,
                    "at_site_bouts": start_count(
                        at_site_intervals, relative_start, relative_stop, include_stop=is_final_bin
                    ),
                    "foraging_bouts": start_count(
                        foraging_intervals, relative_start, relative_stop, include_stop=is_final_bin
                    ),
                    "focal_visible_episodes": start_count(
                        focal_visible_union, relative_start, relative_stop, include_stop=is_final_bin
                    ),
                    "fly_events": point_counts["fly_events"],
                    "jump_events": point_counts["jump_events"],
                    "out_of_sight_events": point_counts["out_of_sight_events"],
                    "displacement_events": point_counts["displacement_events"],
                    "displacement_focal_to_others": point_counts["displacement_focal_to_others"],
                    "displacement_others_to_focal": point_counts["displacement_others_to_focal"],
                    "displacement_unspecified": point_counts["displacement_unspecified"],
                    "subject_count_update_events": point_counts["subject_count_update_events"],
                    "garbage_marker_events": point_counts["garbage_marker_events"],
                }
            )
            current_clock_start = current_clock_stop

        raw_at_site_seconds = interval_length(at_site_intervals)
        raw_foraging_seconds = interval_length(foraging_intervals)
        focal_visible_seconds = interval_length(focal_visible_union)
        qc_rows.append(
            {
                "observation_id": observation_id,
                "at_site_bouts": len(at_site_intervals),
                "foraging_bouts": len(foraging_intervals),
                "raw_at_site_seconds": raw_at_site_seconds,
                "foraging_seconds": raw_foraging_seconds,
                "focal_visible_union_seconds": focal_visible_seconds,
                "added_visible_seconds": extra_visible_seconds,
                "foraging_intervals_outside_at_site": len(containment_failures),
                "at_site_overlap_count": same_behavior_overlap_count(at_site_intervals),
                "foraging_overlap_count": same_behavior_overlap_count(foraging_intervals),
                "negative_duration_count": invalid["negative_duration"],
                "zero_duration_count": invalid["zero_duration"],
                "outside_media_count": invalid["outside_media"],
                "qc_decision": "Union repair applied" if extra_visible_seconds > 0 else "No repair needed",
            }
        )
        day_rows.append(
            {
                "observation_id": observation_id,
                "observation_date": study_date.isoformat(),
                "weekday": study_date.strftime("%A"),
                "observer": observer,
                "collection_day": int(collection_day),
                "media_duration_seconds_source": media_duration,
                "camera_start_datetime": iso_datetime(study_date, camera_start_clock),
                "camera_end_datetime": iso_datetime(study_date, camera_end_clock),
                "collection_end_datetime": iso_datetime(study_date, collection_end_clock)
                if collection_day
                else None,
                "garbage_marker_count": len(garbage_times),
            }
        )

    interval_rows.sort(key=lambda row: (row["observation_date"], row["bin_start_datetime"], row["observation_id"]))
    qc_rows.sort(key=lambda row: row["observation_id"])
    day_rows.sort(key=lambda row: (row["observation_date"], row["observation_id"]))

    expected = {
        "observation_days": len(day_rows),
        "collection_days": sum(row["collection_day"] for row in day_rows),
        "noncollection_days": sum(not row["collection_day"] for row in day_rows),
        "interval_rows": len(interval_rows),
        "media_duration_seconds": sum(row["media_duration_seconds_source"] for row in day_rows),
        "raw_at_site_seconds": sum(row["raw_at_site_seconds"] for row in qc_rows),
        "foraging_seconds": sum(row["foraging_seconds"] for row in qc_rows),
        "focal_visible_seconds": sum(row["focal_visible_union_seconds"] for row in qc_rows),
        "added_visible_seconds": sum(row["added_visible_seconds"] for row in qc_rows),
        "foraging_intervals_outside_at_site": sum(
            row["foraging_intervals_outside_at_site"] for row in qc_rows
        ),
        "at_site_bouts": sum(row["at_site_bouts"] for row in qc_rows),
        "foraging_bouts": sum(row["foraging_bouts"] for row in qc_rows),
        "displacement_events": sum(row["displacement_events"] for row in interval_rows),
        "garbage_marker_events": sum(row["garbage_marker_events"] for row in interval_rows),
    }

    result = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "parameters": {
            "bin_seconds": BIN_SECONDS,
            "containment_tolerance_seconds": CONTAINMENT_TOLERANCE_SECONDS,
            "focal_visible_rule": "Strict union of At the observation site and Searching for food state intervals",
        },
        "sources": [
            {
                "path": str(BORIS_PATH.relative_to(ROOT)),
                "sha256": sha256(BORIS_PATH),
                "role": "Authoritative event and media-duration source",
            },
            {
                "path": str(OBS_TIME_PATH.relative_to(ROOT)),
                "sha256": sha256(OBS_TIME_PATH),
                "role": "Authoritative study date and camera-start source",
            },
            {
                "path": str(AGG_ROOT.relative_to(ROOT)),
                "sha256": None,
                "role": "BORIS state-interval exports used to pair state starts and stops",
            },
        ],
        "expected_totals": expected,
        "day_rows": day_rows,
        "qc_rows": qc_rows,
        "interval_rows": interval_rows,
    }

    OUTPUT_PATH.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(json.dumps(expected, indent=2))


# --- Command-line entry point ---
if __name__ == "__main__":
    main()
